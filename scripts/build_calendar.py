#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_calendar.py — 交易年历 Excel -> docs/calendar.json 转换脚本

用途：
    读取「交易年历.xlsx」第一个 sheet（Daily_Notices），
    把 A 列(日期) / B 列(交易所) / C 列(提示事项) / D 列(合约)
    抽取为前端可直接读取的 JSON，供网页按"当天日期"筛选展示。

数据更新流程（后期维护）：
    1. 在仓库 data/ 目录删除旧的 交易年历.xlsx，上传新的同名文件
       （或直接在 GitHub 网页端 Upload 覆盖）。
    2. 点网页上的「手动刷新」按钮，或等每日定时任务自动运行，
       本脚本会重新生成 docs/calendar.json 并重新部署 Pages。

注意：GitHub Actions 的 cron 运行在 UTC，已在 core/date_utils 用北京时间处理；
本脚本只负责把 Excel 转成 JSON，不关心"今天"的判定（由前端按浏览器北京时间筛选）。
"""

import os
import sys
import json
from datetime import datetime, timezone, timedelta

import openpyxl

# 项目根目录（脚本位于 scripts/ 下，上级即项目根）
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "交易年历.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "docs", "calendar.json")

# 第一个 sheet 名（按需求固定为 Daily_Notices）
SHEET_NAME = "Daily_Notices"

# 标题行（真正表头所在行）。第1行是总标题，第2行空，第3行是字段标题
HEADER_ROW = 3
DATA_START_ROW = 4

# 字段映射：Excel 列 -> JSON key -> 中文标题
FIELD_MAP = {
    "date": 1,      # A 列 日期
    "exchange": 2,  # B 列 交易所
    "matter": 3,    # C 列 提示事项
    "contract": 4,  # D 列 合约
}
HEADERS_CN = {
    "date": "日期",
    "exchange": "交易所",
    "matter": "提示事项",
    "contract": "合约",
}


def to_iso_date(value):
    """把 Excel 单元格值规范化为 YYYY-MM-DD 字符串，无法识别返回 None。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, __import__("datetime").date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] 未找到年历文件: {EXCEL_PATH}")
        print("请先把「交易年历.xlsx」放到仓库的 data/ 目录下。")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] 工作簿中找不到 sheet: {SHEET_NAME}，现有: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # 处理 A 列合并单元格：合并区域的值只存在于左上角单元格，
    # 需要把该值扩散到区域内的每一行，否则同一日期的后续行会丢失日期。
    date_by_row = {}
    for row in range(DATA_START_ROW, ws.max_row + 1):
        date_by_row[row] = to_iso_date(ws.cell(row=row, column=FIELD_MAP["date"]).value)

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col <= FIELD_MAP["date"] <= max_col:
            top_left_value = ws.cell(row=min_row, column=FIELD_MAP["date"]).value
            top_left_date = to_iso_date(top_left_value)
            if top_left_date is not None:
                for row in range(min_row, max_row + 1):
                    date_by_row[row] = top_left_date

    events = []
    skipped = 0
    last_date = None
    last_exchange = ""
    for row in range(DATA_START_ROW, ws.max_row + 1):
        date_iso = date_by_row.get(row)
        if date_iso is None:
            skipped += 1
            continue  # 跳过空行 / 非日期行
        exchange = (ws.cell(row=row, column=FIELD_MAP["exchange"]).value or "").strip()
        matter = (ws.cell(row=row, column=FIELD_MAP["matter"]).value or "").strip()
        contract = (ws.cell(row=row, column=FIELD_MAP["contract"]).value or "").strip()
        # 跳过完全空白的行
        if not exchange and not matter and not contract:
            skipped += 1
            continue
        # 同一日期块内，B 列为空时继承上一个非空交易所
        if not exchange and date_iso == last_date and last_exchange:
            exchange = last_exchange
        if exchange:
            last_exchange = exchange
        last_date = date_iso
        events.append({
            "date": date_iso,
            "exchange": exchange,
            "matter": matter,
            "contract": contract,
        })

    # 按日期升序，同一日期内保持原始行顺序
    events.sort(key=lambda e: e["date"])

    BJT = timezone(timedelta(hours=8))
    payload = {
        "generated_at": datetime.now(BJT).isoformat(),
        "source_file": os.path.basename(EXCEL_PATH),
        "source_sheet": SHEET_NAME,
        "headers": HEADERS_CN,
        "event_count": len(events),
        "events": events,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"[OK] 已生成 {OUT_PATH}")
    print(f"     事件总数: {len(events)}")
    print(f"     日期范围: {events[0]['date']} ~ {events[-1]['date']}")


if __name__ == "__main__":
    main()
